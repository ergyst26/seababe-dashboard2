from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
from jose import JWTError, jwt
from passlib.context import CryptContext
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import aiofiles
from openpyxl import Workbook
from io import BytesIO

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
SECRET_KEY = os.environ.get('JWT_SECRET', 'orderflow-secret-key-2024')
ALGORITHM = "HS256"

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# File upload directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Create the main app
app = FastAPI()

# Security
security = HTTPBearer()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ============ MODELS ============

class UserCreate(BaseModel):
    email: str
    password: str
    name: str

class UserLogin(BaseModel):
    email: str
    password: str

class ClientCreate(BaseModel):
    name: str
    surname: str
    ig_name: str = ""
    address: str = ""
    phone: str = ""
    photo: str = ""

class ClientUpdate(BaseModel):
    name: Optional[str] = None
    surname: Optional[str] = None
    ig_name: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    photo: Optional[str] = None

class OrderCreate(BaseModel):
    client_id: str
    total_price: float
    shipping_type: str = "paid"
    product_photo: str = ""
    masa: str = ""
    notes: str = ""

class OrderUpdate(BaseModel):
    client_id: Optional[str] = None
    total_price: Optional[float] = None
    shipping_type: Optional[str] = None
    product_photo: Optional[str] = None
    masa: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None

# ============ AUTH HELPERS ============

def create_access_token(data: dict):
    to_encode = data.copy()
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("user_id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token i pavlefshëm")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="Përdoruesi nuk u gjet")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Token i pavlefshëm")

# ============ AUTH ROUTES ============

@api_router.post("/auth/register")
async def register(user: UserCreate):
    # Check max 2 users
    user_count = await db.users.count_documents({})
    if user_count >= 2:
        raise HTTPException(status_code=400, detail="Numri maksimal i përdoruesve është arritur (2)")
    
    # Check if email exists
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Ky email ekziston tashmë")
    
    hashed_password = pwd_context.hash(user.password)
    # First user is admin, second is user
    role = "admin" if user_count == 0 else "user"
    user_doc = {
        "id": str(uuid.uuid4()),
        "email": user.email,
        "password": hashed_password,
        "name": user.name,
        "role": role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_access_token({"user_id": user_doc["id"]})
    return {"token": token, "user": {"id": user_doc["id"], "email": user_doc["email"], "name": user_doc["name"], "role": user_doc["role"]}}

@api_router.post("/auth/login")
async def login(user: UserLogin):
    db_user = await db.users.find_one({"email": user.email}, {"_id": 0})
    if not db_user or not pwd_context.verify(user.password, db_user["password"]):
        raise HTTPException(status_code=401, detail="Email ose fjalëkalimi i gabuar")
    
    token = create_access_token({"user_id": db_user["id"]})
    return {"token": token, "user": {"id": db_user["id"], "email": db_user["email"], "name": db_user["name"], "role": db_user.get("role", "admin")}}

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    return {"id": user["id"], "email": user["email"], "name": user["name"], "role": user.get("role", "admin")}

# ============ CLIENT ROUTES ============

@api_router.get("/clients")
async def get_clients(user=Depends(get_current_user)):
    clients = await db.clients.find({}, {"_id": 0}).sort("created_at", 1).to_list(1000)
    return clients

@api_router.post("/clients")
async def create_client(client_data: ClientCreate, user=Depends(get_current_user)):
    client_doc = {
        "id": str(uuid.uuid4()),
        "name": client_data.name,
        "surname": client_data.surname,
        "ig_name": client_data.ig_name,
        "address": client_data.address,
        "phone": client_data.phone,
        "photo": client_data.photo,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.clients.insert_one(client_doc)
    return {k: v for k, v in client_doc.items() if k != "_id"}

@api_router.put("/clients/{client_id}")
async def update_client(client_id: str, client_data: ClientUpdate, user=Depends(get_current_user)):
    update_dict = {k: v for k, v in client_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="Asnjë fushë për të përditësuar")
    
    result = await db.clients.update_one({"id": client_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Klienti nuk u gjet")
    
    updated = await db.clients.find_one({"id": client_id}, {"_id": 0})
    return updated

@api_router.delete("/clients/{client_id}")
async def delete_client(client_id: str, user=Depends(get_current_user)):
    result = await db.clients.delete_one({"id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Klienti nuk u gjet")
    return {"message": "Klienti u fshi me sukses"}

# ============ UPLOAD ROUTE ============

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), user=Depends(get_current_user)):
    file_ext = Path(file.filename).suffix
    file_name = f"{uuid.uuid4()}{file_ext}"
    file_path = UPLOAD_DIR / file_name
    
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    return {"filename": file_name, "url": f"/api/uploads/{file_name}"}

# ============ ORDER ROUTES ============

@api_router.get("/orders")
async def get_orders(user=Depends(get_current_user)):
    pipeline = [
        {"$sort": {"created_at": 1}},
        {"$lookup": {
            "from": "clients",
            "localField": "client_id",
            "foreignField": "id",
            "as": "client"
        }},
        {"$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}},
        {"$addFields": {
            "client_name": {
                "$cond": {
                    "if": "$client",
                    "then": {"$concat": ["$client.name", " ", "$client.surname"]},
                    "else": "I panjohur"
                }
            },
            "client_ig": {"$ifNull": ["$client.ig_name", ""]}
        }},
        {"$project": {"_id": 0, "client": 0}}
    ]
    orders = await db.orders.aggregate(pipeline).to_list(1000)
    return orders

@api_router.post("/orders")
async def create_order(order_data: OrderCreate, user=Depends(get_current_user)):
    # Verify client exists
    client = await db.clients.find_one({"id": order_data.client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Klienti nuk u gjet")
    
    order_doc = {
        "id": str(uuid.uuid4()),
        "client_id": order_data.client_id,
        "total_price": order_data.total_price,
        "shipping_type": order_data.shipping_type,
        "product_photo": order_data.product_photo,
        "masa": order_data.masa,
        "notes": order_data.notes,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.orders.insert_one(order_doc)
    result = {k: v for k, v in order_doc.items() if k != "_id"}
    result["client_name"] = f"{client['name']} {client['surname']}"
    result["client_ig"] = client.get("ig_name", "")
    return result

@api_router.put("/orders/{order_id}")
async def update_order(order_id: str, order_data: OrderUpdate, user=Depends(get_current_user)):
    update_dict = {k: v for k, v in order_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="Asnjë fushë për të përditësuar")
    
    result = await db.orders.update_one({"id": order_id}, {"$set": update_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Porosia nuk u gjet")
    
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    # Enrich with client info
    client = await db.clients.find_one({"id": updated.get("client_id")}, {"_id": 0})
    if client:
        updated["client_name"] = f"{client['name']} {client['surname']}"
        updated["client_ig"] = client.get("ig_name", "")
    return updated

@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str, user=Depends(get_current_user)):
    # Only admin can delete orders
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Vetëm admini mund të fshijë porosi")
    result = await db.orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Porosia nuk u gjet")
    return {"message": "Porosia u fshi me sukses"}

# ============ EXPORT ORDERS TO EXCEL ============

@api_router.get("/orders/export")
async def export_orders(
    start_date: str = Query(..., description="Start date YYYY-MM-DD"),
    end_date: str = Query(..., description="End date YYYY-MM-DD"),
    user=Depends(get_current_user)
):
    # Parse dates
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formati i datës i gabuar. Përdorni YYYY-MM-DD")

    # Get orders in date range using aggregation
    pipeline = [
        {"$match": {"created_at": {"$gte": start.isoformat(), "$lte": end.isoformat()}}},
        {"$sort": {"created_at": 1}},
        {"$lookup": {
            "from": "clients",
            "localField": "client_id",
            "foreignField": "id",
            "as": "client"
        }},
        {"$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}},
        {"$addFields": {
            "client_name": {
                "$cond": {
                    "if": "$client",
                    "then": {"$concat": ["$client.name", " ", "$client.surname"]},
                    "else": "I panjohur"
                }
            },
            "client_ig": {"$ifNull": ["$client.ig_name", ""]}
        }},
        {"$project": {"_id": 0, "client": 0}}
    ]
    orders = await db.orders.aggregate(pipeline).to_list(10000)

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Porositë"

    # Headers
    headers = ["IG Username", "Emri i Klientit", "Masa", "Çmimi Total", "Statusi", "Shënime", "Data"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = cell.font.copy(bold=True)

    # Data rows
    for row_idx, order in enumerate(orders, 2):
        ws.cell(row=row_idx, column=1, value=f"@{order.get('client_ig', '')}" if order.get('client_ig') else "")
        ws.cell(row=row_idx, column=2, value=order.get("client_name", ""))
        ws.cell(row=row_idx, column=3, value=order.get("masa", ""))
        ws.cell(row=row_idx, column=4, value=order.get("total_price", 0))
        status = "E Përfunduar" if order.get("status") == "completed" else "Në Pritje"
        ws.cell(row=row_idx, column=5, value=status)
        ws.cell(row=row_idx, column=6, value=order.get("notes", ""))
        ws.cell(row=row_idx, column=7, value=order.get("created_at", "")[:10])

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"porosi_{start_date}_deri_{end_date}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ DASHBOARD STATS ============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user=Depends(get_current_user)):
    total_clients = await db.clients.count_documents({})
    total_orders = await db.orders.count_documents({})
    pending_orders = await db.orders.count_documents({"status": "pending"})
    completed_orders = await db.orders.count_documents({"status": "completed"})
    
    # Total revenue
    pipeline = [{"$group": {"_id": None, "total": {"$sum": "$total_price"}}}]
    revenue_result = await db.orders.aggregate(pipeline).to_list(1)
    total_revenue = revenue_result[0]["total"] if revenue_result else 0
    
    # Recent orders with $lookup
    recent_pipeline = [
        {"$sort": {"created_at": -1}},
        {"$limit": 5},
        {"$lookup": {
            "from": "clients",
            "localField": "client_id",
            "foreignField": "id",
            "as": "client"
        }},
        {"$unwind": {"path": "$client", "preserveNullAndEmptyArrays": True}},
        {"$addFields": {
            "client_name": {
                "$cond": {
                    "if": "$client",
                    "then": {"$concat": ["$client.name", " ", "$client.surname"]},
                    "else": "I panjohur"
                }
            }
        }},
        {"$project": {"_id": 0, "client": 0}}
    ]
    recent_orders = await db.orders.aggregate(recent_pipeline).to_list(5)
    
    return {
        "total_clients": total_clients,
        "total_orders": total_orders,
        "pending_orders": pending_orders,
        "completed_orders": completed_orders,
        "total_revenue": total_revenue,
        "recent_orders": recent_orders
    }

# Include the router in the main app
app.include_router(api_router)

# Serve uploaded files
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
